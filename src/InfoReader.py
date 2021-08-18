from openpyxl import load_workbook

from src.Student import Student


class InfoReader:
    """
    读取信息的工具类
    """

    @staticmethod
    def get_student_info(filename):
        """
            :arg
                filename:文件名称，如"学生个人博客信息.xlsx"
            :returns
                学生的列表
        """
        workbook = load_workbook('../src/text/' + filename)
        sheets = workbook.get_sheet_names()
        booksheet = workbook.get_sheet_by_name(sheets[0])
        rows = booksheet.rows
        i = 1
        students = list()
        for row in rows:
            id = booksheet.cell(row=i, column=1).value
            name = booksheet.cell(row=i, column=2).value
            url = booksheet.cell(row=i, column=3).value
            if id is None:
                break
            id = str(id)
            student = Student(id=id[0:10], name=name, url=url)
            students.append(student)
            i = i + 1
        return students
