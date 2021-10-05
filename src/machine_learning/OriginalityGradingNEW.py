import os
import random
import re

import cv2
import numpy as np
import tensorflow as tf
from keras.callbacks import ModelCheckpoint, ReduceLROnPlateau
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import KFold, train_test_split
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import classification_report
from sklearn.svm import SVC

from src.machine_learning import data_analysis, model_generator
from src.InfoReadAndWrite import InfoReadAndWrite
from src.tools.SplitDataset import SplitDataset


class OriginalityGradingNEW:
    """
    博客原创性分级类(1000篇文章)
    """

    @staticmethod
    def get_labels_xlsx(filepath: str, num: int) -> list:
        """
        从xlsx文件中读取标签
        :param filepath: 文件路径
        :param num: 标签数量
        :return: 标签
        """
        workbook = load_workbook(filepath)
        sheets = workbook.get_sheet_names()
        booksheet = workbook.get_sheet_by_name(sheets[0])
        rows = booksheet.rows
        i = 1
        labels = list()
        for row in rows:
            if i > num:
                break
            label = booksheet.cell(row=i, column=2).value
            labels.append(int(label))
            i = i + 1
        return labels

    @staticmethod
    def get_labels(filepath: str, num: int) -> list:
        """
        从txt文件中读取标签
        :param filepath: 文件路径
        :param num: 标签数量
        :return: 标签
        """
        i = 1
        labels = list()
        with open(filepath, "r") as f:
            for line in f.readlines():
                if i > num:
                    break
                i += 1
                line = re.sub("[\\t ]+", "\t", line)
                label = int(line.split("\t")[2].replace("\n", ""))
                labels.append(label)
        return labels

    @staticmethod
    def change_labels(labels):
        """
        更改一些label
        :param labels: 原来的labels
        :return: 更改后的labels
        """
        new_labels = list(labels)
        with open('../../text/差距大的文章.txt', 'r') as f:
            for line in f.readlines():
                line = re.sub('\\t+', '\\t', line)
                line = line.split('\t')
                if len(line) >= 4:
                    num = int(line[0])
                    label = int(line[3].replace('\n', ''))
                    new_labels[num] = label
        return new_labels

    @staticmethod
    def img_show(vector):
        """
        将一个1320维的向量展示为一张灰度图
        :param vector: 1320维的向量，其值在0-1之间
        :return:
        """
        im_array = np.expand_dims(np.reshape(vector, [60, 37]), axis=-1)
        cv2.imshow(" ", cv2.resize(im_array, (2400, 1480)))
        cv2.waitKey(0)

    @staticmethod
    def data_show(vector):
        """
        展示向量的意义
        :param vector: 向量
        :return: 无
        """
        print('标题相似度')
        print(vector[:10])
        print('全文相似度')
        print(vector[10:20])
        print('段落相似度')
        for i in range(20, 820, 10):
            print('{}>>>'.format((i - 20) / 10), end='')
            print(vector[i: i + 10])
        print('句子相似度')
        for i in range(820, 1620, 10):
            print('{}>>>'.format((i - 820) / 10), end='')
            print(vector[i: i + 10])
        print('代码相似度')
        for i in range(1620, 2220, 10):
            print('{}>>>'.format((i - 1620) / 10), end='')
            print(vector[i: i + 10])

    @staticmethod
    def reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len):
        print(vectors.shape)
        new_vectors = []
        for vector in vectors:
            new_vector = []
            new_vector.extend(vector[:similarity_len])
            new_vector.extend(vector[10:10 + similarity_len])
            for i in range(20, 20 + paragraph_len * 10, 10):
                new_vector.extend(vector[i:i + similarity_len])
            for i in range(820, 820 + sentence_len * 10, 10):
                new_vector.extend(vector[i:i + similarity_len])
            for i in range(1620, 1620 + code_len * 10, 10):
                new_vector.extend(vector[i:i + similarity_len])
            new_vectors.append(new_vector)
        vectors = np.float32(new_vectors)
        print('-------转换之后的size-------')
        print(vectors.shape)
        return vectors

    @staticmethod
    def get_results(y_true, y_pred, test_index=None, label_filepath=None, filepath=None):
        """
        获得标签的差异值
        :param label_filepath: 存储标签的文件
        :param filepath: 需要保存的路径
        :param test_index: 序号
        :param y_true: 真实值
        :param y_pred: 预测值
        :return: 无
        """
        dictionary = dict()
        abs_result = []
        if test_index and label_filepath and filepath:
            with open(label_filepath, "r") as f:
                for line in f.readlines():
                    line = re.sub("[\\t ]+", "\t", line)
                    line = line.split("\t")
                    if len(line) >= 3:
                        dictionary[int(line[0])] = line[1]
        result = []
        for i in range(len(y_true)):
            sub = abs(int(y_pred[i]) - int(y_true[i]))
            if sub != 0:
                abs_result.append(sub)
                if test_index and label_filepath and filepath:
                    result.append('{}\t{}\t\t{}\t{}'.format(test_index[i], dictionary[test_index[i]], int(y_true[i]),
                                                            int(y_pred[i])))
        if test_index and label_filepath and filepath:
            if os.path.exists(filepath):
                print("\n" + filepath + ">>>已存在\n", end="")
                return 1
            with open(filepath, "w") as f:
                for r in result:
                    f.write(r)
                    f.write('\n')
        return abs_result

    @staticmethod
    def reduce_dimension(vectors, head: bool, text: bool, paragraphs: bool, sentences: bool, codes: bool):
        """
        减少维度
        :param vectors: 原矩阵
        :param head: 是否不需要head
        :param text: 是否不需要text
        :param paragraphs: 是否不需要paragraphs
        :param sentences: 是否不需要sentences
        :param codes: 是否不需要codes
        :return: 新矩阵
        """
        new_vectors = []
        for vector in vectors:
            new_vector = []
            if head:
                new_vector.extend(vector[:10])
            if text:
                new_vector.extend(vector[10:20])
            if paragraphs:
                new_vector.extend(vector[20:820])
            if sentences:
                new_vector.extend(vector[820:1620])
            if codes:
                new_vector.extend(vector[1620:])
            new_vectors.append(new_vector)
        return np.float32(new_vectors)

    @staticmethod
    def one_dimension(vectors, mode: str):
        """
        将每个句子等的n个相似度取max,min,mean,median。
        :param vectors: 源矩阵
        :param mode: 模式。max：取最大值；min：取最小值；mean：取平均值；median。
        :return: 新矩阵
        """
        new_vectors = []
        for vector in vectors:
            new_vector = []
            for i in range(0, 2220, 10):
                if mode == 'max':
                    new_vector.append(np.max(vector[i:i + 10]))
                if mode == 'min':
                    new_vector.append(np.min(vector[i:i + 10]))
                if mode == 'mean':
                    new_vector.append(np.mean(vector[i:i + 10]))
                if mode == 'median':
                    new_vector.append(np.median(vector[i:i + 10]))
            new_vectors.append(new_vector)
        return np.float32(new_vectors)

    @staticmethod
    def random_select(vectors, dimension: int):
        """
        随机选取一些维度
        :param vectors: 原矩阵
        :param dimension: 期望维度
        :return: 新矩阵
        """
        new_vectors = []
        for vector in vectors:
            sample_list = [i for i in range(len(vector))]
            sample_list = random.sample(sample_list, dimension)
            new_vector = vector[sample_list]
            new_vectors.append(new_vector)
        return np.float32(new_vectors)

    @staticmethod
    def low_k(vectors, k: int, code_start: int, k_code: int):
        """
        选取相似度最小的k+k_code个相似度
        :param k_code: 代码取多少k_code个
        :param code_start: 代码段段开始下标
        :param vectors: 原矩阵
        :param k: 非码取多少k个
        :return: 新矩阵
        """
        new_vectors = []
        for vector in vectors:
            new_vectors.append(sorted(vector[:code_start])[:k] + sorted(vector[code_start:])[:k_code])
        return np.float32(new_vectors)

    @staticmethod
    def loose_report(y_true: [], y_predict: []):
        """
        将0,1,2,3,4共5个分类的准确评判标准改为如下：
        实际为0->预测为0,1
        实际为1->预测为0,1,2
        实际为2->预测为1,2,3
        实际为3->预测为2,3,4
        实际为4->预测为3,4
        并且打印结果
        :param y_true: 真实值
        :param y_predict: 预测值
        :return: 结果报告
        """
        y_predict_new = list(y_predict)
        for i in range(len(y_predict)):
            if y_true[i] == 0 and (y_predict[i] == 0 or y_predict[i] == 1):
                y_predict_new[i] = 0
            if y_true[i] == 1 and (y_predict[i] == 0 or y_predict[i] == 1 or y_predict[i] == 2):
                y_predict_new[i] = 1
            elif y_true[i] == 2 and (y_predict[i] == 1 or y_predict[i] == 2 or y_predict[i] == 3):
                y_predict_new[i] = 2
            elif y_true[i] == 3 and (y_predict[i] == 2 or y_predict[i] == 3 or y_predict[i] == 4):
                y_predict_new[i] = 3
            elif y_predict[i] == 4 and (y_predict[i] == 3 or y_predict[i] == 4):
                y_predict_new[i] = 4
        return classification_report(y_true, y_predict_new)

    @staticmethod
    def merge_labels(labels: []) -> ([[]], []):
        """
        将标签合并，规则如下：
        0->0
        1,2,3->1
        4->2
        :param labels: 原标签
        :return: 新标签
        """
        new_labels = []
        for i in range(len(labels)):
            if labels[i] == 0:
                new_labels.append(0)
            if labels[i] == 1 or labels[i] == 2 or labels[i] == 3:
                new_labels.append(1)
            if labels[i] == 4:
                new_labels.append(2)
        return new_labels

    @staticmethod
    def split_result_show(labels, model, test_index, vectors):
        sp = SplitDataset('../../text/按原创性分类.txt')
        data_set = sp.data_split(test_index, language_rate=0.9, length=200, code_rate=0.3)
        if data_set['chinese']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['chinese']], ), axis=-1))
            report = classification_report(labels[data_set['chinese']], y_predict)
            print('<<<chinese>>>')
            print(report)
        if data_set['english']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['english']], ), axis=-1))
            report = classification_report(labels[data_set['english']], y_predict)
            print('<<<english>>>')
            print(report)
        if data_set['mixed']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['mixed']], ), axis=-1))
            report = classification_report(labels[data_set['mixed']], y_predict)
            print('<<<mixed>>>')
            print(report)
        if data_set['long']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['long']], ), axis=-1))
            report = classification_report(labels[data_set['long']], y_predict)
            print('<<<long>>>')
            print(report)
        if data_set['short']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['short']], ), axis=-1))
            report = classification_report(labels[data_set['short']], y_predict)
            print('<<<short>>>')
            print(report)
        if data_set['text']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['text']], ), axis=-1))
            report = classification_report(labels[data_set['text']], y_predict)
            print('<<<text>>>')
            print(report)
        if data_set['code']:
            y_predict = list(tf.argmax(model.predict(vectors[data_set['code']], ), axis=-1))
            report = classification_report(labels[data_set['code']], y_predict)
            print('<<<code>>>')
            print(report)
        multi_topics, single_topic = sp.one_more_topics(test_index)
        y_predict = list(tf.argmax(model.predict(vectors[multi_topics], ), axis=-1))
        report = classification_report(labels[multi_topics], y_predict)
        print('<<<multi_topics>>>')
        print(report)
        y_predict = list(tf.argmax(model.predict(vectors[single_topic], ), axis=-1))
        report = classification_report(labels[single_topic], y_predict)
        print('<<<single_topics>>>')
        print(report)


def traditional_ml(data_filepath, label_filepath):
    n_splits = 5
    random_state = 40

    labels = OriginalityGradingNEW.get_labels(label_filepath, 650)
    labels = OriginalityGradingNEW.change_labels(labels)
    labels = OriginalityGradingNEW.merge_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)
    k_fold = KFold(n_splits=n_splits, random_state=random_state, shuffle=True)
    total_y_predict_knn = []
    total_y_predict_DT = []
    total_y_predict_Bayes = []
    total_y_predict_SVM = []
    total_y_predict_RF = []
    total_y_predict_clf = []
    total_y_test = []
    for train_index, test_index in k_fold.split(vectors, labels):
        x_train, x_test, y_train, y_test = vectors[train_index], vectors[test_index], labels[train_index], labels[
            test_index]
        total_y_test.extend(y_test)

        dt = DecisionTreeClassifier().fit(x_train, y_train)
        answer_dt = dt.predict(x_test)
        total_y_predict_DT.extend(answer_dt)

        gnb = GaussianNB().fit(x_train, y_train)
        answer_gnb = gnb.predict(x_test)
        total_y_predict_Bayes.extend(answer_gnb)

        clf = LogisticRegression(penalty='l1', solver='liblinear').fit(x_train, y_train)
        answer_clf = clf.predict(x_test)
        total_y_predict_clf.extend(answer_clf)

        # tuned_parameters = [{'kernel': ['rbf'], 'gamma': [1e-1, 1e-2, 1e-3], 'C': [1, 10, 100, 1000]},
        #                     {'kernel': ['linear'], 'C': [1, 10, 100, 1000]}]
        #
        # scores = ['precision', 'recall']
        # for score in scores:
        #     print("# Tuning hyper-parameters for %s" % score)
        #     print('%s_weighted' % score)
        #     grid = GridSearchCV(SVC(), tuned_parameters, cv=5,
        #                         scoring='%s_weighted' % score)  # 1/5作为验证集
        #     # 用前一半train数据再做5折交叉验证，因为之前train_test_split已经分割为2份了
        #
        #     grid_result = grid.fit(x_train, y_train)
        #
        #     print("Best parameters set found on development set:")
        #     print(grid.best_params_)
        #     print("Grid scores on development set:")
        #
        #     means = grid_result.cv_results_['mean_test_score']
        #     params = grid_result.cv_results_['params']
        #     for mean, param in zip(means, params):
        #         print("%f  with:   %r" % (mean, param))
        #
        #     print("Detailed classification report:")
        #     print("The model is trained on the full development set.")
        #     print("The scores are computed on the full evaluation set.")
        #     y_true, y_pred = y_test, grid.predict(x_test)
        #     print(classification_report(y_true, y_pred))
        #
        # tuned_parameters = [{'n_estimators': list(range(10, 210, 10))}]
        # scores = ['precision', 'recall']
        # for score in scores:
        #     print("# Tuning hyper-parameters for %s" % score)
        #     print('%s_weighted' % score)
        #     grid = GridSearchCV(RandomForestClassifier(), tuned_parameters, cv=5,
        #                         scoring='%s_weighted' % score)  # 1/5作为验证集
        #     # 用前一半train数据再做5折交叉验证，因为之前train_test_split已经分割为2份了
        #
        #     grid_result = grid.fit(x_train, y_train)
        #
        #     print("Best parameters set found on development set:")
        #     print(grid.best_params_)
        #     print("Grid scores on development set:")
        #
        #     means = grid_result.cv_results_['mean_test_score']
        #     params = grid_result.cv_results_['params']
        #     for mean, param in zip(means, params):
        #         print("%f  with:   %r" % (mean, param))
        #
        #     print("Detailed classification report:")
        #     print("The model is trained on the full development set.")
        #     print("The scores are computed on the full evaluation set.")
        #     y_true, y_pred = y_test, grid.predict(x_test)
        #     print(classification_report(y_true, y_pred))
        #
        # tuned_parameters = [{'n_neighbors': list(range(1, 20))}]
        # scores = ['precision', 'recall']
        # for score in scores:
        #     print("# Tuning hyper-parameters for %s" % score)
        #     print('%s_weighted' % score)
        #     grid = GridSearchCV(KNeighborsClassifier(), tuned_parameters, cv=5,
        #                         scoring='%s_weighted' % score)  # 1/5作为验证集
        #     # 用前一半train数据再做5折交叉验证，因为之前train_test_split已经分割为2份了
        #
        #     grid_result = grid.fit(x_train, y_train)
        #
        #     print("Best parameters set found on development set:")
        #     print(grid.best_params_)
        #     print("Grid scores on development set:")
        #
        #     means = grid_result.cv_results_['mean_test_score']
        #     params = grid_result.cv_results_['params']
        #     for mean, param in zip(means, params):
        #         print("%f  with:   %r" % (mean, param))
        #
        #     print("Detailed classification report:")
        #     print("The model is trained on the full development set.")
        #     print("The scores are computed on the full evaluation set.")
        #     y_true, y_pred = y_test, grid.predict(x_test)
        #     print(classification_report(y_true, y_pred))
        # break

        svm = SVC(C=10, gamma=0.01, kernel='rbf', probability=True).fit(x_train, y_train)
        answer_svm = svm.predict(x_test)
        total_y_predict_SVM.extend(answer_svm)

        rf = RandomForestClassifier(n_estimators=80).fit(x_train, y_train)
        answer_rf = rf.predict(x_test)
        total_y_predict_RF.extend(answer_rf)

        knn = KNeighborsClassifier(n_neighbors=19).fit(x_train, y_train)
        answer_knn = knn.predict(x_test)
        total_y_predict_knn.extend(answer_knn)

        # print('\n\nThe classification report for knn:')
        # print(classification_report(y_test, answer_knn))
        # print('\n\nThe classification report for DT:')
        # print(classification_report(y_test, answer_dt))
        # print('\n\nThe classification report for Bayes:')
        # print(classification_report(y_test, answer_gnb))
        # print('\n\nThe classification report for SVM:')
        # print(classification_report(y_test, answer_svm))
    print("总结果如下")
    print('\n\nThe classification report for knn:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_knn))
    print(classification_report(total_y_test, total_y_predict_knn))
    print('\n\nThe classification report for DT:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_DT))
    print(classification_report(total_y_test, total_y_predict_DT))
    print('\n\nThe classification report for Bayes:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_Bayes))
    print(classification_report(total_y_test, total_y_predict_Bayes))
    print('\n\nThe classification report for SVM:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_SVM))
    print(classification_report(total_y_test, total_y_predict_SVM))
    print('\n\nThe classification report for RF:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_RF))
    print(classification_report(total_y_test, total_y_predict_RF))
    print('\n\nThe classification report for CLF:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_clf))
    print(classification_report(total_y_test, total_y_predict_clf))


def traditional_ml_new(data_filepath, label_filepath):
    n_splits = 5
    random_state = 40

    labels = OriginalityGradingNEW.get_labels_xlsx(label_filepath, 650)
    labels = OriginalityGradingNEW.merge_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)
    k_fold = KFold(n_splits=n_splits, random_state=random_state, shuffle=True)
    total_y_predict_knn = []
    total_y_predict_DT = []
    total_y_predict_Bayes = []
    total_y_predict_SVM = []
    total_y_predict_RF = []
    total_y_predict_clf = []
    total_y_test = []
    for train_index, test_index in k_fold.split(vectors, labels):
        x_train, x_test, y_train, y_test = vectors[train_index], vectors[test_index], labels[train_index], labels[
            test_index]
        total_y_test.extend(y_test)

        dt = DecisionTreeClassifier().fit(x_train, y_train)
        answer_dt = dt.predict(x_test)
        total_y_predict_DT.extend(answer_dt)

        gnb = GaussianNB().fit(x_train, y_train)
        answer_gnb = gnb.predict(x_test)
        total_y_predict_Bayes.extend(answer_gnb)

        clf = LogisticRegression(penalty='l1', solver='liblinear').fit(x_train, y_train)
        answer_clf = clf.predict(x_test)
        total_y_predict_clf.extend(answer_clf)

        # tuned_parameters = [{'kernel': ['rbf'], 'gamma': [1e-1, 1e-2, 1e-3], 'C': [1, 10, 100, 1000]},
        #                     {'kernel': ['linear'], 'C': [1, 10, 100, 1000]}]
        # scores = ['precision', 'recall']
        # for score in scores:
        #     print("# Tuning hyper-parameters for %s" % score)
        #     print('%s_weighted' % score)
        #     grid = GridSearchCV(SVC(), tuned_parameters, cv=5,
        #                         scoring='%s_weighted' % score)  # 1/5作为验证集
        #     # 用前一半train数据再做5折交叉验证，因为之前train_test_split已经分割为2份了
        #
        #     grid_result = grid.fit(x_train, y_train)
        #
        #     print("Best parameters set found on development set:")
        #     print(grid.best_params_)
        #     print("Grid scores on development set:")
        #
        #     means = grid_result.cv_results_['mean_test_score']
        #     params = grid_result.cv_results_['params']
        #     for mean, param in zip(means, params):
        #         print("%f  with:   %r" % (mean, param))
        #
        #     print("Detailed classification report:")
        #     print("The model is trained on the full development set.")
        #     print("The scores are computed on the full evaluation set.")
        #     y_true, y_pred = y_test, grid.predict(x_test)
        #     print(classification_report(y_true, y_pred))
        #
        # tuned_parameters = [{'n_estimators': list(range(10, 210, 10))}]
        # scores = ['precision', 'recall']
        # for score in scores:
        #     print("# Tuning hyper-parameters for %s" % score)
        #     print('%s_weighted' % score)
        #     grid = GridSearchCV(RandomForestClassifier(), tuned_parameters, cv=5,
        #                         scoring='%s_weighted' % score)  # 1/5作为验证集
        #     # 用前一半train数据再做5折交叉验证，因为之前train_test_split已经分割为2份了
        #
        #     grid_result = grid.fit(x_train, y_train)
        #
        #     print("Best parameters set found on development set:")
        #     print(grid.best_params_)
        #     print("Grid scores on development set:")
        #
        #     means = grid_result.cv_results_['mean_test_score']
        #     params = grid_result.cv_results_['params']
        #     for mean, param in zip(means, params):
        #         print("%f  with:   %r" % (mean, param))
        #
        #     print("Detailed classification report:")
        #     print("The model is trained on the full development set.")
        #     print("The scores are computed on the full evaluation set.")
        #     y_true, y_pred = y_test, grid.predict(x_test)
        #     print(classification_report(y_true, y_pred))
        #
        # tuned_parameters = [{'n_neighbors': list(range(1, 30))}]
        # scores = ['precision', 'recall']
        # for score in scores:
        #     print("# Tuning hyper-parameters for %s" % score)
        #     print('%s_weighted' % score)
        #     grid = GridSearchCV(KNeighborsClassifier(), tuned_parameters, cv=5,
        #                         scoring='%s_weighted' % score)  # 1/5作为验证集
        #     # 用前一半train数据再做5折交叉验证，因为之前train_test_split已经分割为2份了
        #
        #     grid_result = grid.fit(x_train, y_train)
        #
        #     print("Best parameters set found on development set:")
        #     print(grid.best_params_)
        #     print("Grid scores on development set:")
        #
        #     means = grid_result.cv_results_['mean_test_score']
        #     params = grid_result.cv_results_['params']
        #     for mean, param in zip(means, params):
        #         print("%f  with:   %r" % (mean, param))
        #
        #     print("Detailed classification report:")
        #     print("The model is trained on the full development set.")
        #     print("The scores are computed on the full evaluation set.")
        #     y_true, y_pred = y_test, grid.predict(x_test)
        #     print(classification_report(y_true, y_pred))
        # break

        svm = SVC(C=1, gamma=0.01, kernel='rbf', probability=True).fit(x_train, y_train)
        answer_svm = svm.predict(x_test)
        total_y_predict_SVM.extend(answer_svm)

        rf = RandomForestClassifier(n_estimators=190).fit(x_train, y_train)
        answer_rf = rf.predict(x_test)
        total_y_predict_RF.extend(answer_rf)

        knn = KNeighborsClassifier(n_neighbors=23).fit(x_train, y_train)
        answer_knn = knn.predict(x_test)
        total_y_predict_knn.extend(answer_knn)

        # print('\n\nThe classification report for knn:')
        # print(classification_report(y_test, answer_knn))
        # print('\n\nThe classification report for DT:')
        # print(classification_report(y_test, answer_dt))
        # print('\n\nThe classification report for Bayes:')
        # print(classification_report(y_test, answer_gnb))
        # print('\n\nThe classification report for SVM:')
        # print(classification_report(y_test, answer_svm))
    print("总结果如下")
    print('\n\nThe classification report for knn:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_knn))
    print(classification_report(total_y_test, total_y_predict_knn))
    print('\n\nThe classification report for DT:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_DT))
    print(classification_report(total_y_test, total_y_predict_DT))
    print('\n\nThe classification report for Bayes:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_Bayes))
    print(classification_report(total_y_test, total_y_predict_Bayes))
    print('\n\nThe classification report for SVM:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_SVM))
    print(classification_report(total_y_test, total_y_predict_SVM))
    print('\n\nThe classification report for RF:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_RF))
    print(classification_report(total_y_test, total_y_predict_RF))
    print('\n\nThe classification report for CLF:')
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict_clf))
    print(classification_report(total_y_test, total_y_predict_clf))


def train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, get_model, embedding_len,
          drop_out_rate, learning_rate, l1, l2, verbose):
    k_fold = KFold(n_splits=n_splits, random_state=random_state, shuffle=True)
    total_y_predict = []
    total_y_test = []
    vectors, x_valid, labels, y_valid = train_test_split(vectors, labels, test_size=0.2, shuffle=True,
                                                         random_state=random_state)
    y_valid = np.int32(tf.keras.utils.to_categorical(y_valid, num_classes=nd))
    for train_index, test_index in k_fold.split(vectors, labels):
        model = get_model(embedding_len, drop_out_rate, learning_rate, l1, l2, nd)
        x_train, x_test, y_train, y_test = vectors[train_index], vectors[test_index], labels[train_index], labels[
            test_index]
        y_train = np.int32(tf.keras.utils.to_categorical(y_train, num_classes=nd))

        # history = model.fit(x_train, y_train, epochs=epochs, verbose=verbose, batch_size=batch_size,
        #                     validation_split=0.2,
        #                     validation_freq=validation_freq, callbacks=callbacks, shuffle=True)
        history = model.fit(x_train, y_train, epochs=epochs, verbose=verbose, batch_size=batch_size,
                            validation_data=(x_valid, y_valid), callbacks=callbacks, shuffle=True)

        data_analysis.show_history(history, is_accuracy=True)
        # model = tf.keras.models.load_model(filepath)
        # print(model.predict(x_train))
        print(model.evaluate(x_train, y_train))
        y_predict = list(tf.argmax(model.predict(x_test), axis=-1))
        print(classification_report(y_test, y_predict))
        # print(OriginalityGradingNEW.loose_report(y_test, y_predict))
        data_analysis.plot_confusion_matrix(y_test, y_predict, [0, 1, 2, 3, 4])
        # OriginalityGradingNEW.get_results(y_test, y_predict, test_index, label_filepath,
        #                                   filepath='../src/text/差距大的文章3.txt')
        total_y_predict.extend(y_predict)
        total_y_test.extend(y_test)
        # split_result_show(labels, model, test_index, vectors)
        # return 0
    print("总结果")
    print(classification_report(total_y_test, total_y_predict))
    # print(OriginalityGradingNEW.loose_report(total_y_test, total_y_predict))
    data_analysis.plot_confusion_matrix(total_y_test, total_y_predict, [0, 1, 2, 3, 4])
    print(OriginalityGradingNEW.get_results(total_y_test, total_y_predict))


def dense(data_filepath, label_filepath):
    verbose = 1
    nd = 3
    filepath = "../saved_model/originality_grading_model.h5"
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', min_delta=0,
                                                      patience=100, verbose=verbose, mode='auto',
                                                      baseline=None, restore_best_weights=False)
    check_pointer = ModelCheckpoint(filepath=filepath, verbose=verbose, save_best_only=True, monitor='accuracy')
    learning_rate_reduction = ReduceLROnPlateau(monitor='loss', patience=5, verbose=verbose,
                                                factor=0.5, min_lr=1e-5)
    callbacks = []
    labels = OriginalityGradingNEW.get_labels(label_filepath, 650)
    labels = OriginalityGradingNEW.change_labels(labels)
    labels = OriginalityGradingNEW.merge_labels(labels)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)

    # vectors, labels = OriginalityGradingNEW.merge_labels(vectors, labels)
    # nd = 4

    labels = np.array(labels, dtype=int)
    vectors = np.array(vectors, dtype=float)

    embedding_len = 2220
    # random_state = 1  # 测试
    random_state = 40  # 一直使用的
    validation_freq = 1
    n_splits = 5
    drop_out_rate = 0.05
    l1 = 0
    l2 = 0.1
    # callbacks.append(early_stopping)
    # callbacks.append(check_pointer)
    callbacks.append(learning_rate_reduction)

    similarity_len = 10
    paragraph_len = 50
    sentence_len = 50
    code_len = 50
    # vectors = OriginalityGradingNEW.reduce_dimension(vectors,
    #                                                  head=True, text=True, paragraphs=True, sentences=True, codes=False)
    # embedding_len = 1620
    # vectors = OriginalityGradingNEW.reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len)
    # vectors = OriginalityGradingNEW.random_select(vectors, 100)
    # vectors = OriginalityGradingNEW.one_dimension(vectors, 'max')
    # k = 20
    # k_code = 5
    # vectors = OriginalityGradingNEW.low_k(vectors, k, 162, k_code)
    # print(vectors.shape)
    # print(vectors)
    # return 0

    learning_rate = 1e-1
    batch_size = 256
    # epochs = 500
    epochs = 500
    # 总的准确率0.82
    train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, model_generator.get_model_dense,
          embedding_len, drop_out_rate, learning_rate, l1, l2, verbose)


def cnn(data_filepath, label_filepath):
    verbose = 1
    nd = 5
    filepath = "../saved_model/originality_grading_model.h5"
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', min_delta=0,
                                                      patience=100, verbose=verbose, mode='auto',
                                                      baseline=None, restore_best_weights=False)
    check_pointer = ModelCheckpoint(filepath=filepath, verbose=verbose, save_best_only=True, monitor='accuracy')
    learning_rate_reduction = ReduceLROnPlateau(monitor='loss', patience=10, verbose=verbose,
                                                factor=0.5, min_lr=1e-5)
    callbacks = []
    labels = OriginalityGradingNEW.get_labels(label_filepath, 650)
    labels = OriginalityGradingNEW.change_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)
    reshaped_vectors = []
    for x_ in vectors:
        reshaped_vectors.append(tf.expand_dims(tf.reshape(x_, [60, 37]), axis=-1))
    vectors = np.array(reshaped_vectors)

    embedding_len = 2220
    random_state = 40  # 0.72
    validation_freq = 1
    n_splits = 5
    drop_out_rate = 0.2
    l1 = 0.01
    l2 = 0.01
    # callbacks.append(early_stopping)
    # callbacks.append(check_pointer)
    callbacks.append(learning_rate_reduction)

    # similarity_len = 10
    # paragraph_len = 60
    # sentence_len = 60
    # code_len = 10
    # vectors = OriginalityGradingNEW.reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len)

    learning_rate = 1e-1
    # batch_size = 6000,9000 # 0.75
    batch_size = 32
    epochs = 2000
    # 总的准确率0.71
    train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, model_generator.get_model_cnn,
          embedding_len, drop_out_rate, learning_rate, l1, l2, verbose)


def rnn(data_filepath, label_filepath):
    verbose = 2
    nd = 5
    filepath = "../saved_model/originality_grading_model.h5"
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', min_delta=0,
                                                      patience=100, verbose=verbose, mode='auto',
                                                      baseline=None, restore_best_weights=False)
    check_pointer = ModelCheckpoint(filepath=filepath, verbose=verbose, save_best_only=True, monitor='accuracy')
    callbacks = []
    labels = OriginalityGradingNEW.get_labels(label_filepath, 650)
    labels = OriginalityGradingNEW.change_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)

    embedding_len = 2220
    # random_state = 30  # 0.71
    random_state = 40  # 0.72
    validation_freq = 1
    n_splits = 5
    drop_out_rate = 0.3
    l1 = 0.01
    l2 = 0.01
    # callbacks.append(early_stopping)
    # callbacks.append(check_pointer)

    # similarity_len = 10
    # paragraph_len = 60
    # sentence_len = 60
    # code_len = 10
    # vectors = OriginalityGradingNEW.reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len)

    learning_rate = 1e-1
    # batch_size = 6000,9000 # 0.75
    batch_size = 3000
    epochs = 200
    # 总的准确率0.71
    train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, model_generator.get_model_rnn,
          embedding_len, drop_out_rate, learning_rate, l1, l2, verbose)


def lstm(data_filepath, label_filepath):
    verbose = 2
    nd = 5
    filepath = "../saved_model/originality_grading_model.h5"
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', min_delta=0,
                                                      patience=100, verbose=verbose, mode='auto',
                                                      baseline=None, restore_best_weights=False)
    check_pointer = ModelCheckpoint(filepath=filepath, verbose=verbose, save_best_only=True, monitor='accuracy')
    callbacks = []
    labels = OriginalityGradingNEW.get_labels(label_filepath, 650)
    labels = OriginalityGradingNEW.change_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)

    embedding_len = 2220
    # random_state = 30  # 0.71
    random_state = 40  # 0.72
    validation_freq = 1
    n_splits = 5
    drop_out_rate = 0.3
    l1 = 0.01
    l2 = 0.01
    # callbacks.append(early_stopping)
    # callbacks.append(check_pointer)

    # similarity_len = 10
    # paragraph_len = 60
    # sentence_len = 60
    # code_len = 10
    # vectors = OriginalityGradingNEW.reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len)

    learning_rate = 1e-1
    # batch_size = 6000,9000 # 0.75
    batch_size = 3000
    epochs = 200
    # 总的准确率0.71
    train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, model_generator.get_model_lstm,
          embedding_len, drop_out_rate, learning_rate, l1, l2, verbose)


def gru(data_filepath, label_filepath):
    verbose = 2
    nd = 5
    filepath = "../saved_model/originality_grading_model.h5"
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', min_delta=0,
                                                      patience=100, verbose=verbose, mode='auto',
                                                      baseline=None, restore_best_weights=False)
    check_pointer = ModelCheckpoint(filepath=filepath, verbose=verbose, save_best_only=True, monitor='accuracy')
    callbacks = []
    labels = OriginalityGradingNEW.get_labels(label_filepath, 650)
    labels = OriginalityGradingNEW.change_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)

    embedding_len = 2220
    # random_state = 30  # 0.71
    random_state = 40  # 0.72
    validation_freq = 1
    n_splits = 5
    drop_out_rate = 0.3
    l1 = 0.01
    l2 = 0.01
    # callbacks.append(early_stopping)
    # callbacks.append(check_pointer)

    # similarity_len = 10
    # paragraph_len = 60
    # sentence_len = 60
    # code_len = 10
    # vectors = OriginalityGradingNEW.reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len)

    learning_rate = 1e-3
    # batch_size = 6000,9000 # 0.75
    batch_size = 3000
    epochs = 200
    # 总的准确率0.71
    model = model_generator.get_model_gru(embedding_len, drop_out_rate, learning_rate, l1, l2, nd)
    train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, model_generator.get_model_gru,
          embedding_len, drop_out_rate, learning_rate, l1, l2, verbose)


def dense_new(data_filepath, label_filepath):
    verbose = 1
    nd = 3
    filepath = "../saved_model/originality_grading_model.h5"
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', min_delta=0,
                                                      patience=100, verbose=verbose, mode='auto',
                                                      baseline=None, restore_best_weights=False)
    check_pointer = ModelCheckpoint(filepath=filepath, verbose=verbose, save_best_only=True, monitor='accuracy')
    learning_rate_reduction = ReduceLROnPlateau(monitor='loss', patience=5, verbose=verbose,
                                                factor=0.5, min_lr=1e-5)
    callbacks = []
    labels = OriginalityGradingNEW.get_labels_xlsx(label_filepath, 650)
    labels = OriginalityGradingNEW.merge_labels(labels)
    labels = np.array(labels, dtype=int)
    vectors = InfoReadAndWrite.get_similarities(data_filepath)

    embedding_len = 2220
    # random_state = 1  # 测试
    random_state = 40  # 一直使用的
    n_splits = 5
    drop_out_rate = 0.05
    l1 = 0
    l2 = 0.1
    # callbacks.append(early_stopping)
    # callbacks.append(check_pointer)
    callbacks.append(learning_rate_reduction)

    similarity_len = 10
    paragraph_len = 50
    sentence_len = 50
    code_len = 50
    # vectors = OriginalityGradingNEW.reduce_dimension(vectors,
    #                                                  head=True, text=True, paragraphs=True, sentences=True, codes=False)
    # embedding_len = 1620
    # vectors = OriginalityGradingNEW.reshape(vectors, similarity_len, paragraph_len, sentence_len, code_len)
    # vectors = OriginalityGradingNEW.random_select(vectors, 100)
    # vectors = OriginalityGradingNEW.one_dimension(vectors, 'max')
    # k = 20
    # k_code = 5
    # vectors = OriginalityGradingNEW.low_k(vectors, k, 162, k_code)
    # print(vectors.shape)
    # print(vectors)
    # return 0

    learning_rate = 1e-1
    batch_size = 256
    epochs = 500
    # 总的准确率0.82
    train(batch_size, callbacks, epochs, labels, n_splits, nd, random_state, vectors, model_generator.get_model_dense,
          embedding_len, drop_out_rate, learning_rate, l1, l2, verbose)


if __name__ == '__main__':
    data_filepath = "../text/similarities_bigger.csv"
    data_filepath_new = '../text/similarities_bigger_new.csv'
    label_filepath = "../text/按原创性分类.txt"
    label_filepath_new = "../text/按原创性分类_改.xlsx"

    # 旧的数据
    # traditional_ml(data_filepath, label_filepath)  # 已完成，旧的标签
    # traditional_ml_new(data_filepath, label_filepath_new)  # 已完成，新的标签
    # dense(data_filepath, label_filepath)  # 已完成，旧的标签
    # dense_new(data_filepath, label_filepath_new)  # 已完成，新的标签

    # 新的的数据
    # traditional_ml(data_filepath_new, label_filepath)  # 已完成，旧的标签
    # traditional_ml_new(data_filepath_new, label_filepath_new)  # 已完成，新的标签
    # dense(data_filepath_new, label_filepath)  # 已完成，旧的标签
    dense_new(data_filepath_new, label_filepath_new)  # 已完成，新的标签

    # cnn(data_filepath, label_filepath)
    # gru(data_filepath, label_filepath)
    # rnn(data_filepath, label_filepath)
    # lstm(data_filepath, label_filepath)

    # vectors = InfoReadAndWrite.get_similarities(data_filepath)
    # i = 649
    # OriginalityGradingNEW.data_show(vectors[i])
    # OriginalityGradingNEW.img_show(vectors[i])
