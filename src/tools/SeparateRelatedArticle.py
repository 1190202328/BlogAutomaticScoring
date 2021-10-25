import json
import re

import numpy as np
from keras.callbacks import ReduceLROnPlateau
from openpyxl import load_workbook
from sklearn.metrics import classification_report
from sklearn.model_selection import KFold, train_test_split
import tensorflow as tf

from src.machine_learning import data_analysis
from src.tools import Clean, GetWebResource


class SeparateRelatedArticle:
    """
    将文章列表分离出软件构造主题相关文章的工具类
    """

    @staticmethod
    def get_urls(filepath):
        """
        从磁盘中读如urls
        :param filepath: 文件路径
        :return: urls列表
        """
        urls = list()
        f = open(filepath, "r")
        for line in f.readlines():
            r = line.split("\t")
            urls.append(r[1].replace("\n", " "))
        f.close()
        return urls

    @staticmethod
    def write_article_to_file(urls, filepath):
        """
        将urls列表对应的文章都写入文件
        :param filepath:
        :return:
        """
        results = list()
        i = 1
        for url in urls:
            print("正在处理第{}个url(共{}个)>>>".format(i, len(urls)) + url)
            i += 1
            results.append(GetWebResource.split_txt(url, EDU=False, verbose=False))
        with open(filepath, 'w') as f:
            f.write(json.dumps(results, ensure_ascii=False, indent=2))

    @staticmethod
    def get_results(filepath):
        """
        获得根据urls得到的results
        :return: results
        """
        with open(filepath, 'r') as f:
            return json.loads(f.read())

    @staticmethod
    def get_sequences(texts, embedding_len, vocab_list=None):
        """
        获取句子对应的词序列
        :param texts: 文章列表
        :param embedding_len: 每个向量维度
        :param vocab_list: 词典列表
        :return: 词序列
        """
        if not vocab_list:
            vocab_list = list()
            f = open("../../text/vocab_list_text.txt", 'r')
            for line in f.readlines():
                vocab_list.append(line[:-1])
            f.close()
        token_list = list()
        for text in texts:
            i = 0
            token = list()
            for word in text:
                if i > embedding_len:
                    break
                i += 1
                if word not in vocab_list:
                    token.append(1)
                else:
                    token.append(vocab_list.index(word))
            token = token[:embedding_len] + [0] * (embedding_len - len(token))
            token_list.append(token)
        token_list = np.array(token_list)
        return token_list

    @staticmethod
    def get_texts_and_labels():
        texts = list()
        labels = list()
        source = open("../../text/texts.txt", mode="r")
        pattern = re.compile("\\d+")
        for line in source.readlines():
            if re.match("第(\\d+)篇文章\\[\\d*\\]", line):
                result = pattern.findall(line)
                if len(result) == 1:
                    labels.append(0)
                else:
                    if int(result[1]) == 0:
                        labels.append(0)
                    else:
                        labels.append(1)
            else:
                texts.append(line)
        return texts, labels

    @staticmethod
    def get_course_related_urls(urls):
        vocab_list = list()
        f = open("../../text/vocab_list_text.txt", 'r')
        for line in f.readlines():
            vocab_list.append(line[:-1])
        f.close()
        valid_urls = list()
        course_related_urls = list()
        texts = list()
        embedding_len = 200
        count = 1
        for url in urls:
            print("[{}]>>>{}".format(count, url))
            count += 1
            result = GetWebResource.split_txt(url, False, verbose=False)
            if result and result.get('text'):
                valid_urls.append(url)
                texts.append(result.get('text'))
        texts = Clean.clean_with_low_frequency(texts)
        sequences = SeparateRelatedArticle.get_sequences(texts, embedding_len, vocab_list=vocab_list)
        model = tf.keras.models.load_model("../saved_model/related_text_separate_model.h5")
        y_pred = model.predict(sequences)
        for i in range(len(y_pred)):
            if y_pred[i][0] > y_pred[i][1]:
                course_related_urls.append(valid_urls[i])
        return course_related_urls


def machine_learning():
    texts, labels = SeparateRelatedArticle.get_texts_and_labels()
    labels = np.array(labels)
    texts = np.array(Clean.clean_with_low_frequency(texts), dtype=object)
    random_state = 40
    texts, x_valid, labels, y_valid = train_test_split(texts, labels, test_size=0.2, shuffle=True,
                                                       random_state=random_state)
    k_fold = KFold(n_splits=5, random_state=random_state, shuffle=True)
    for train_index, test_index in k_fold.split(texts, labels):
        x_train, x_test, y_train, y_test = texts[train_index], texts[test_index], labels[train_index], labels[
            test_index]
        vocab = set()
        for x in x_train:
            for word in x:
                vocab.add(word)
        # 深度学习分类
        vocab_list = list()
        vocab_list.append("<paddle>")
        vocab_list.append("<unk>")
        vocab_list += list(sorted(vocab))
        # f = open("../../text/vocab_list_text_1.txt", 'w')
        # for vocab in vocab_list:
        #     f.write(vocab)
        #     f.write("\n")
        # f.close()
        embedding_len = 100
        output_dim = 64
        batch_size = 128
        epochs = 500
        verbose = 1
        learning_rate = 1e-1
        l1 = 0
        l2 = 1e-2
        drop_out_rate = 1e-2
        vocab_len = len(vocab_list)
        opt = tf.optimizers.Adadelta(learning_rate=learning_rate)
        learning_rate_reduction = ReduceLROnPlateau(monitor='loss', patience=10, verbose=verbose,
                                                    factor=0.5, min_lr=1e-5)

        x_train = SeparateRelatedArticle.get_sequences(x_train, embedding_len, vocab_list)
        x_test = SeparateRelatedArticle.get_sequences(x_test, embedding_len, vocab_list)
        x_valid_sequences = SeparateRelatedArticle.get_sequences(x_valid, embedding_len, vocab_list)

        model = tf.keras.Sequential([
            tf.keras.layers.InputLayer(input_shape=(embedding_len,)),
            tf.keras.layers.Embedding(input_dim=vocab_len, output_dim=output_dim),
            tf.keras.layers.BatchNormalization(),
            tf.keras.layers.ActivityRegularization(l1=l1, l2=l2),
            tf.keras.layers.Bidirectional(tf.keras.layers.GRU(output_dim)),
            tf.keras.layers.BatchNormalization(),
            tf.keras.layers.Dense(32, activation=tf.nn.relu),
            tf.keras.layers.Dropout(drop_out_rate),
            tf.keras.layers.Dense(2, activation=tf.nn.softmax)
        ])
        model.compile(optimizer=opt, loss=tf.keras.losses.sparse_categorical_crossentropy, metrics=['accuracy'])

        print(model.summary())
        # history = model.fit(x_train, y_train, epochs=epochs, validation_data=(x_valid_sequences, y_valid),
        #                     verbose=verbose,
        #                     batch_size=batch_size,
        #                     callbacks=[learning_rate_reduction])
        # data_analysis.show_history(history, is_accuracy=True)

        path = "../../src/saved_model/"
        filename = "related_text_separate_model_1.h5"
        # model.save(path + filename)
        model = tf.keras.models.load_model(path + filename)
        y_predict = list(tf.argmax(model.predict(x_test), axis=-1))
        print(classification_report(y_test, y_predict))
        break


def get_total_urls():
    workbook = load_workbook("../../text/blog.xlsx")
    sheets = workbook.get_sheet_names()
    urls = list()
    for sheet in sheets:
        booksheet = workbook.get_sheet_by_name(sheet)
        rows = booksheet.rows
        i = 2
        for _ in rows:
            url = booksheet.cell(row=i, column=1).value
            i += 1
            if url is None or url == "":
                continue
            url = re.sub("\\s+", "", str(url))
            if not re.match("(https://.*)|(http://.*)", url):
                url = "https://" + url
            if re.match('https://=HYPERLINK.*', url):
                url = url.split('"')[1]
            urls.append(url)
    return urls


def refresh_index():
    lines = list()
    with open('../../text/按原创性分类.txt', mode='r') as f:
        i = 0
        for line in f.readlines():
            lines.append(str(i) + line[line.find('\t'):])
            i += 1
    print(len(lines))
    f1 = open('../src/text/按原创性分类的更新.txt', mode='w')
    for line in lines:
        f1.write(line)
    f1.close()


if __name__ == '__main__':
    machine_learning()

    # total_urls_get = []
    # with open('../../text/所有学生的所有url.txt', mode='r') as f:
    #     for line in f.readlines():
    #         total_urls_get.append(line[:-1])
    # print(len(total_urls_get))
    #
    # total_urls = []
    # urls = get_total_urls()
    # urls += total_urls_get
    # print(len(urls))
    # i = 1
    # for url in urls:
    #     main_url = GetWebResource.get_main_url(url)
    #     if main_url is not None and main_url != "":
    #         print('<{}> '.format(i), main_url)
    #         i += 1
    #         local_urls = GetWebResource.get_urls(main_url, verbose=False)
    #         total_urls += local_urls
    #
    # print(len(total_urls))
    # with open('../../text/所有文章的url.txt', mode='w') as f:
    #     for url in total_urls:
    #         f.write(url)
    #         f.write('\n')
    #
    # total_urls_get = []
    # with open('../../text/所有文章的url.txt', mode='r') as f:
    #     for line in f.readlines():
    #         total_urls_get.append(line[:-1])
    # print(len(total_urls_get))

    # course_related_urls = SeparateRelatedArticle.get_course_related_urls(total_urls)
    # with open('../src/text/按原创性分类.txt', "w") as f:
    #     i = 0
    #     for url in course_related_urls:
    #         f.write('{}\t{}\n'.format(i, url))
    #         i += 1

    # refresh_index()
